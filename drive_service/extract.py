"""
drive_service/extract.py
========================
Turn any downloaded file into plain text. Deliberately dumb — it does NOT try to
understand the data, it just flattens it to text. The LLM in compare-core does
the understanding, which is what makes this reusable across payroll / census /
contracts / financial data with no per-domain parsers.

Supports: csv, tsv, xlsx, pdf, txt/everything-else (read as utf-8).
"""

import csv
import io
import os


def _ext(path):
    return os.path.splitext(path)[1].lower().lstrip(".")


def _sniff(path):
    """Detect format from the file's first bytes, for files saved without an
    extension (e.g. Drive downloads cached by file id). Returns an ext string."""
    with open(path, "rb") as f:
        head = f.read(8)
    if head[:4] == b"PK\x03\x04":      # zip container -> xlsx/xlsm
        return "xlsx"
    if head[:4] == b"%PDF":            # pdf
        return "pdf"
    return "csv"                       # default: treat as delimited/plain text


def _from_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except Exception:
            dialect = csv.excel
        for row in csv.reader(f, dialect):
            rows.append(", ".join(c.strip() for c in row))
    return "\n".join(rows)


def _from_xlsx(path):
    from openpyxl import load_workbook
    # Read bytes and wrap in BytesIO so openpyxl doesn't reject by extension
    # (Drive downloads are cached as the bare file id, no .xlsx suffix).
    with open(path, "rb") as f:
        data = io.BytesIO(f.read())
    wb = load_workbook(data, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                out.append(", ".join(cells))
    wb.close()
    return "\n".join(out)


def _from_pdf(path):
    try:
        import pdfplumber
        chunks = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                for tbl in page.extract_tables() or []:
                    for row in tbl:
                        chunks.append(", ".join("" if c is None else str(c) for c in row))
                txt = page.extract_text() or ""
                if txt.strip():
                    chunks.append(txt)
        return "\n".join(chunks)
    except Exception:
        with open(path, "rb") as f:
            return f.read().decode("utf-8", errors="replace")


def to_text(path):
    """Read a file and return plain text. Never raises on format — falls back to raw read."""
    ext = _ext(path) or _sniff(path)
    if ext not in ("csv", "tsv", "xlsx", "xlsm", "pdf"):
        ext = _sniff(path)   # unknown extension -> trust the bytes
    try:
        if ext in ("csv", "tsv"):
            return _from_csv(path)
        if ext in ("xlsx", "xlsm"):
            return _from_xlsx(path)
        if ext == "pdf":
            return _from_pdf(path)
    except Exception as e:
        print(f"  extract: {ext} parse failed for {path} ({e}); falling back to raw read")
    with open(path, "rb") as f:
        return f.read().decode("utf-8", errors="replace")
